import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import hashlib

LOG_DIR = "/mnt/logs"
os.makedirs(LOG_DIR, exist_ok=True)

def carregar_mapeamento_codigos(path):
    tabela = {}
    try:
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo, sku = row
            if codigo is not None and sku is not None:
                tabela[round(float(codigo), 1)] = str(sku)
    except Exception as e:
        print(f"Erro ao carregar mapeamento: {e}")
    return tabela

CAMINHO_MAPA_SKU = os.path.join(os.path.dirname(__file__), "SKU.xlsx")
MAPA_SKU = carregar_mapeamento_codigos(CAMINHO_MAPA_SKU)

primario_valor = {0: 1.0, 1: 2.0, 2: 3.0, 3: 4.0, 4: 5.0, 5: 6.0}
secundario_valor = {0: 0.1, 1: 0.2, 2: 0.3, 5: 0.4}
combinacoes_validas = [(0, 0), (0, 1), (0, 2), (1, 0), (2, 0), (1, 1),
                       (5, 5), (5, 0), (0, 5), (5, 1), (1, 5)]

def gerar_hash_selecao(altura, fio, malha):
    dados = f"{altura}-{fio}-{malha}"
    return hashlib.md5(dados.encode()).hexdigest()

def calcular_saida_personalizado(altura_idx, selecionados, alturas_exibidas):
    fio = selecionados["fio"]
    malha = selecionados["malha"]
    if fio is None or malha is None:
        return None
    base_idx = alturas_exibidas[altura_idx][1]
    alturas_com_base = [i for i in selecionados["altura"] if alturas_exibidas[i][1] == base_idx]
    altura_val = primario_valor.get(base_idx, 0) if len(alturas_com_base) == 1 or altura_idx == alturas_com_base[0] else secundario_valor.get(base_idx, primario_valor.get(base_idx, 0))
    return round(altura_val + 10 * (fio + 1) + 100 * (malha + 1), 1)

def calcular_saida(selecionados, alturas_exibidas):
    altura_sel = selecionados["altura"]
    fio = selecionados["fio"]
    malha = selecionados["malha"]
    if not altura_sel:
        return None, "Selecione ao menos uma altura"
    if len(altura_sel) == 2:
        base1 = alturas_exibidas[altura_sel[0]][1]
        base2 = alturas_exibidas[altura_sel[1]][1]
        if (base1, base2) not in combinacoes_validas:
            return None, "Combinação de alturas inválida"
    if fio is None or malha is None:
        return None, "Complete todas as seleções"
    altura_val = sum(calcular_saida_personalizado(i, selecionados, alturas_exibidas) - 10 * (fio + 1) - 100 * (malha + 1) for i in altura_sel)
    return round(altura_val + 10 * (fio + 1) + 100 * (malha + 1), 1), None

def registrar_log(valor, mapa=MAPA_SKU):
    nome_maq = "S06"
    data_hoje = datetime.now().strftime("%y%m%d")
    caminho = os.path.join(LOG_DIR, f"{nome_maq}{data_hoje}.xlsx")

    sku = mapa.get(valor)
    if not sku:
        print(f"⚠️ Código {valor} não encontrado na planilha de SKUs.")
        return

    if os.path.exists(caminho):
        wb = load_workbook(caminho)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["SKU", "Quantidade"])

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == sku:
            atual = ws.cell(row=row, column=2).value or 0
            ws.cell(row=row, column=2, value=atual + 1)
            break
    else:
        ws.append([sku, 1])

    wb.save(caminho)
