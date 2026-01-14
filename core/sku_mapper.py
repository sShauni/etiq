"""
Mapeamento de códigos numéricos para SKUs.
Responsável por carregar e gerenciar a tabela de conversão.
"""

from openpyxl import load_workbook
from typing import Dict, Optional
import os


class SKUMapper:
    """Gerencia o mapeamento entre códigos numéricos e SKUs."""
    
    def __init__(self, excel_path: str):
        """
        Inicializa o mapper com o caminho do arquivo Excel.
        
        Args:
            excel_path: Caminho para o arquivo SKU.xlsx
        """
        self.excel_path = excel_path
        self._mapping: Dict[float, str] = {}
        self._load_mapping()
    
    def _load_mapping(self) -> None:
        """Carrega o mapeamento do arquivo Excel."""
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Arquivo SKU não encontrado: {self.excel_path}")
        
        try:
            wb = load_workbook(self.excel_path, read_only=True)
            ws = wb.active
            
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                codigo, sku = row
                
                if codigo is not None and sku is not None:
                    # Arredonda para 1 casa decimal para evitar problemas de float
                    codigo_normalizado = round(float(codigo), 1)
                    self._mapping[codigo_normalizado] = str(sku)
                    row_count += 1
            
            wb.close()
            print(f"✓ SKU Mapper carregado: {row_count} mapeamentos")
            
        except Exception as e:
            raise RuntimeError(f"Erro ao carregar mapeamento SKU: {e}")
    
    def get_sku(self, codigo: float) -> Optional[str]:
        """
        Obtém o SKU correspondente a um código numérico.
        
        Args:
            codigo: Código numérico (ex: 111.1)
        
        Returns:
            SKU correspondente ou None se não encontrado
        """
        codigo_normalizado = round(float(codigo), 1)
        return self._mapping.get(codigo_normalizado)
    
    def has_code(self, codigo: float) -> bool:
        """Verifica se um código existe no mapeamento."""
        codigo_normalizado = round(float(codigo), 1)
        return codigo_normalizado in self._mapping
    
    def reload(self) -> None:
        """Recarrega o mapeamento do arquivo (útil se o Excel foi atualizado)."""
        self._mapping.clear()
        self._load_mapping()
    
    def get_all_codes(self) -> list:
        """Retorna lista de todos os códigos mapeados."""
        return sorted(self._mapping.keys())
    
    def get_mapping_count(self) -> int:
        """Retorna o número total de mapeamentos."""
        return len(self._mapping)


if __name__ == '__main__':
    # Teste do mapper
    from config.settings import settings
    
    try:
        mapper = SKUMapper(settings.sku_file_path)
        print(f"Total de SKUs: {mapper.get_mapping_count()}")
        
        # Testa alguns códigos
        test_codes = [111.1, 112.1, 121.1]
        for code in test_codes:
            sku = mapper.get_sku(code)
            print(f"Código {code} -> SKU: {sku if sku else 'NÃO ENCONTRADO'}")
            
    except Exception as e:
        print(f"Erro no teste: {e}")