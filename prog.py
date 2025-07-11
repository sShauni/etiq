import tkinter as tk
from tkinter import messagebox
import os
from datetime import datetime
import subprocess
import threading
import time
import hashlib
from openpyxl import Workbook, load_workbook

LOG_DIR = "/mnt/logs"
os.makedirs(LOG_DIR, exist_ok=True)

def carregar_mapeamento_codigos(path):
    tabela = {}
    try:
        wb = load_workbook(path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo, sku = row
            if codigo is not None and sku is not None:
                tabela[round(float(codigo), 1)] = str(sku)
    except Exception as e:
        print(f"Erro ao carregar mapeamento: {e}")
    return tabela

CAMINHO_MAPA_SKU = os.path.join(os.path.dirname(__file__), "SKU.xlsx")
MAPA_SKU = carregar_mapeamento_codigos(CAMINHO_MAPA_SKU)

def gerar_hash_selecao():
    dados = f"{selecionados['altura']}-{selecionados['fio']}-{selecionados['malha']}"
    return hashlib.md5(dados.encode()).hexdigest()

def registrar_log(valor, automatica=False):
    #nome_maq = "S06"
    data_hoje = datetime.now().strftime("%y%m%d")
    caminho = os.path.join(LOG_DIR, f"{data_hoje}.xlsx")

    sku = MAPA_SKU.get(valor)
    if not sku:
        print(f"⚠️ Código {valor} não encontrado na planilha de SKUs.")
        return

    if os.path.exists(caminho):
        wb = load_workbook(caminho)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["SKU", "Quantidade"])

    encontrou = False
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == sku:
            atual = ws.cell(row=row, column=2).value or 0
            ws.cell(row=row, column=2, value=atual + 1)
            encontrou = True
            break

    if not encontrou:
        ws.append([sku, 1])

    wb.save(caminho)

try:
    import RPi.GPIO as GPIO
    GPIO.setmode(GPIO.BCM)
    PINO_SINAL = 6
    GPIO.setup(PINO_SINAL, GPIO.IN, pull_up_down=GPIO.PUD_DOWN)
    gpio_disponivel = True
except (ImportError, RuntimeError):
    gpio_disponivel = False

MODO_TESTE = True
NOME_IMPRESSORA = "Thermal"
PASTA_ETIQUETAS = os.path.join(os.path.dirname(__file__), "etiquetas")
EXTENSAO_ARQUIVO = ".pdf"
DELAY_IMPRESSAO_GPIO = 1

alturas_exibidas = [
    ("1,00m", 0), ("1,00m (2ª)", 0),
    ("1,20m", 1), ("1,20m (2ª)", 1),
    ("1,50m", 2), ("1,50m (2ª)", 2),
    ("1,80m", 3),
    ("2,00m", 4),
    ("0,50m", 5), ("0,50m (2ª)", 5)
]

botoes_visiveis = [
    True, True,
    True, True,
    True, True,
    True, True,
    True, True


]
fios = ["1,24mm", "1,60mm", "1.90mm", "2.30mm", "2.76mm", "2.10mm"]
malhas = ["5x10cm", "6,5x15cm", "5x15cm", "2,5x2,5cm", "5x5cm", "5x7,5cm"]

fios_visiveis = [True] * len(fios)
#fios_visiveis[0] = False
#fios_visiveis[1] = False
#fios_visiveis[2] = False
#fios_visiveis[3] = False
#fios_visiveis[4] = False
#fios_visiveis[5] = False

malhas_visiveis = [True] * len(malhas)
#malhas_visiveis[0] = False
#malhas_visiveis[1] = False
#malhas_visiveis[2] = False
#malhas_visiveis[3] = False
#malhas_visiveis[4] = False
#malhas_visiveis[5] = False

combinacoes_validas = [
    (0, 0), (0, 1), (0, 2), (1, 0), (2, 0), (1, 1),
    (5, 5), (5, 0), (0, 5), (5, 1), (1, 5)
]

primario_valor = {0: 1.0, 1: 2.0, 2: 3.0, 3: 4.0, 4: 5.0, 5: 6.0}
secundario_valor = {0: 0.1, 1: 0.2, 2: 0.3, 5: 0.4}

selecionados = {"altura": [], "fio": None, "malha": None}

def selecionar(grupo, idx, botoes, visual_indices=None):
    if grupo == "altura":
        if idx in selecionados["altura"]:
            selecionados["altura"].remove(idx)
        elif len(selecionados["altura"]) < 2:
            selecionados["altura"].append(idx)
        else:
            messagebox.showwarning("Erro", "Só é possível selecionar até duas alturas.")
            return
        for b in botoes:
            b.config(bg="lightgray")
        for sel_idx in selecionados["altura"]:
            if sel_idx in visual_indices:
                vis_idx = visual_indices.index(sel_idx)
                botoes[vis_idx].config(bg="mediumpurple1")
    else:
        selecionados[grupo] = idx
        for i, b in enumerate(botoes):
            b.config(bg="mediumpurple1" if i == idx else "lightgray")
    atualizar_saida()

def calcular_saida_personalizado(altura_idx):
    fio = selecionados["fio"]
    malha = selecionados["malha"]
    if fio is None or malha is None:
        return None
    base_idx = alturas_exibidas[altura_idx][1]
    alturas_com_base = [i for i in selecionados["altura"] if alturas_exibidas[i][1] == base_idx]
    altura_val = primario_valor.get(base_idx, 0) if len(alturas_com_base) == 1 or altura_idx == alturas_com_base[0] else secundario_valor.get(base_idx, primario_valor.get(base_idx, 0))
    return round(altura_val + 10 * (fio + 1) + 100 * (malha + 1), 1)

def calcular_saida():
    altura_sel = selecionados["altura"]
    fio = selecionados["fio"]
    malha = selecionados["malha"]
    if not altura_sel:
        return None, "Selecione ao menos uma altura"
    if len(altura_sel) == 2:
        base1 = alturas_exibidas[altura_sel[0]][1]
        base2 = alturas_exibidas[altura_sel[1]][1]
        if (base1, base2) not in combinacoes_validas:
            return None, "Combinação de alturas inválida"
    if fio is None or malha is None:
        return None, "Complete todas as seleções"
    altura_val = sum(calcular_saida_personalizado(i) - 10 * (fio + 1) - 100 * (malha + 1) for i in altura_sel)
    return round(altura_val + 10 * (fio + 1) + 100 * (malha + 1), 1), None

def atualizar_saida():
    valor, erro = calcular_saida()
    saida_var.set(erro if erro else f"Valor de saída: {valor:.1f}")

def imprimir_etiqueta(automatica=False):
    altura_sel = selecionados["altura"]
    if not altura_sel:
        if not automatica:
            messagebox.showerror("Erro", "Selecione ao menos uma altura")
        return
    if len(altura_sel) == 2:
        base1 = alturas_exibidas[altura_sel[0]][1]
        base2 = alturas_exibidas[altura_sel[1]][1]
        if (base1, base2) not in combinacoes_validas:
            if not automatica:
                messagebox.showerror("Erro", "Combinação de alturas inválida")
            return
    arquivos_impressao = []
    for idx in altura_sel:
        valor_individual = calcular_saida_personalizado(idx)
        if valor_individual is None:
            if not automatica:
                messagebox.showerror("Erro", "Complete todas as seleções")
            return
        nome_arquivo = os.path.join(PASTA_ETIQUETAS, f"etiqueta_{valor_individual:.1f}{EXTENSAO_ARQUIVO}")
        if not os.path.exists(nome_arquivo):
            if not automatica:
                messagebox.showerror("Erro", f"Arquivo não encontrado: {nome_arquivo}")
            return
        arquivos_impressao.append((nome_arquivo, valor_individual))
    for nome_arquivo, valor in arquivos_impressao:
        try:
            if not MODO_TESTE:
                subprocess.run(["lp", "-d", NOME_IMPRESSORA, nome_arquivo], check=True)
            registrar_log(valor, automatica)
        except Exception as e:
            if not automatica:
                messagebox.showerror("Erro", f"Falha ao imprimir: {e}")
            return
    if not automatica:
        nomes = "\n".join(n for n, _ in arquivos_impressao)
        messagebox.showinfo("Sucesso", f"Etiqueta(s) {'selecionadas' if MODO_TESTE else 'impressas'}:\n{nomes}")

janela = tk.Tk()
janela.title("Selecionador de Etiquetas")
janela.attributes("-fullscreen", True)
janela.configure(bg="white")
janela.bind("<Escape>", lambda e: janela.destroy())

frame = tk.Frame(janela, bg="white")
frame.pack(pady=10)

def criar_coluna_altura():
    coluna = tk.Frame(frame, bg="white")
    tk.Label(coluna, text="Altura(s)", font=("Arial", 10, "bold"), bg="lightgreen", width=20).grid(row=0, column=0, columnspan=2)
    botoes, visual_indices, linha, coluna_idx = [], [], 1, 0
    for idx, (texto, _) in enumerate(alturas_exibidas):
        if not botoes_visiveis[idx]: continue
        botao = tk.Button(coluna, text=texto, width=12, height=2, font=("Arial", 10), bg="lightgray",
                          command=lambda i=idx: selecionar("altura", i, botoes, visual_indices))
        botao.grid(row=linha, column=coluna_idx, padx=2, pady=2)
        botoes.append(botao)
        visual_indices.append(idx)
        coluna_idx = 0 if coluna_idx >= 1 else 1
        if coluna_idx == 0: linha += 1
    coluna.pack(side=tk.LEFT, padx=5)

def criar_coluna(titulo, opcoes, grupo, visiveis):
    coluna = tk.Frame(frame, bg="white")
    tk.Label(coluna, text=titulo, font=("Arial", 10, "bold"), bg="lightgreen", width=14).pack()
    botoes = []
    for idx, opcao in enumerate(opcoes):
        if not visiveis[idx]: continue
        b = tk.Button(coluna, text=opcao, width=12, height=2, font=("Arial", 10), bg="lightgray",
                      command=lambda i=idx: selecionar(grupo, i, botoes))
        b.pack(pady=2)
        botoes.append(b)
    coluna.pack(side=tk.LEFT, padx=5)

criar_coluna_altura()
criar_coluna("Fio", fios, "fio", fios_visiveis)
criar_coluna("Malha", malhas, "malha", malhas_visiveis)

saida_var = tk.StringVar()
saida_var.set("Faça suas seleções")
tk.Label(janela, textvariable=saida_var, font=("Arial", 12), bg="white").pack(pady=5)

tk.Button(janela, text="Imprimir etiqueta", command=imprimir_etiqueta,
          bg="green", fg="white", font=("Arial", 12, "bold"),
          width=20, height=2).pack(pady=8)

def monitorar_gpio():
    if not gpio_disponivel:
        return
    while True:
        if GPIO.input(PINO_SINAL) == GPIO.HIGH:
            janela.after(0, lambda: imprimir_etiqueta(automatica=True))
            time.sleep(DELAY_IMPRESSAO_GPIO)
            while GPIO.input(PINO_SINAL) == GPIO.HIGH:
                time.sleep(0.05)

if gpio_disponivel:
    threading.Thread(target=monitorar_gpio, daemon=True).start()

janela.mainloop()
